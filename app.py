"""
Football Player Analyzer - MVP
Tracks a selected player through video and counts: passes, tackles, headers, shots.

Flow:
1. POST /upload         → saves video, returns first frame as base64
2. POST /analyze        → starts background analysis (CSRT tracking + Gemini)
3. GET  /progress/{id}  → poll for status / partial results
4. GET  /export/{id}    → download Excel report
"""

import os
import cv2
import json
import base64
import uuid
import asyncio
import threading
import time
from pathlib import Path
from typing import Optional

import numpy as np
from PIL import Image

import google.generativeai as genai

from fastapi import FastAPI, File, UploadFile, HTTPException, BackgroundTasks
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, JSONResponse
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# App setup
# ──────────────────────────────────────────────
app = FastAPI(title="Football Analyzer")

UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

# In-memory session store  {session_id: {...}}
sessions: dict = {}


# ──────────────────────────────────────────────
# Pydantic models
# ──────────────────────────────────────────────
class ValidateKeyRequest(BaseModel):
    api_key: str

class AnalyzeRequest(BaseModel):
    session_id: str
    x: int
    y: int
    width: int
    height: int
    api_key: str


# ──────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────
def frame_to_b64(frame: np.ndarray, quality: int = 75) -> str:
    _, buf = cv2.imencode(".jpg", frame, [cv2.IMWRITE_JPEG_QUALITY, quality])
    return base64.b64encode(buf).decode()


def cv2_to_pil(frame: np.ndarray) -> Image.Image:
    rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    return Image.fromarray(rgb)


def create_tracker():
    """Create CSRT tracker, compatible with different OpenCV versions."""
    for factory in [
        lambda: cv2.TrackerCSRT_create(),
        lambda: cv2.legacy.TrackerCSRT_create(),
        lambda: cv2.TrackerKCF_create(),       # fallback
        lambda: cv2.legacy.TrackerKCF_create(),
    ]:
        try:
            return factory()
        except AttributeError:
            continue
    raise RuntimeError("No compatible OpenCV tracker found. Install opencv-contrib-python.")


def bbox_is_valid(bbox, frame_shape):
    """Check that bbox hasn't drifted outside the frame."""
    h, w = frame_shape[:2]
    x, y, bw, bh = bbox
    return (
        0 <= x < w and
        0 <= y < h and
        bw > 5 and bh > 5 and
        x + bw <= w and
        y + bh <= h
    )


def draw_player_box(frame: np.ndarray, bbox, color=(0, 0, 255), label="PLAYER") -> np.ndarray:
    out = frame.copy()
    x, y, w, h = [int(v) for v in bbox]
    # Thick red rectangle
    cv2.rectangle(out, (x, y), (x + w, y + h), color, 3)
    # Label background
    label_size, _ = cv2.getTextSize(label, cv2.FONT_HERSHEY_SIMPLEX, 0.7, 2)
    lw, lh = label_size
    cv2.rectangle(out, (x, y - lh - 10), (x + lw + 6, y), color, -1)
    cv2.putText(out, label, (x + 3, y - 4),
                cv2.FONT_HERSHEY_SIMPLEX, 0.7, (255, 255, 255), 2)
    # Corner markers
    corner_len = min(w, h) // 4
    for px, py, dx, dy in [(x, y, 1, 1), (x+w, y, -1, 1), (x, y+h, 1, -1), (x+w, y+h, -1, -1)]:
        cv2.line(out, (px, py), (px + dx * corner_len, py), (0, 255, 100), 4)
        cv2.line(out, (px, py), (px, py + dy * corner_len), (0, 255, 100), 4)
    return out


def gemini_analyze_batch(api_key: str, frames_pil: list, timestamps: list,
                         max_retries: int = 4) -> list:
    """
    Send a batch of annotated PIL frames to Gemini Flash.
    Returns list of detected actions [{type, timestamp, confidence}].
    """
    genai.configure(api_key=api_key)
    model = genai.GenerativeModel("gemini-2.0-flash")

    ts_str = ", ".join(f"{t:.1f}s" for t in timestamps)

    prompt = f"""You are a professional football analyst reviewing video frames.

CRITICAL INSTRUCTIONS:
- The player you must analyze is highlighted with a RED BOUNDING BOX labeled "PLAYER"
- You must ONLY evaluate actions performed by the player INSIDE the red box
- Completely IGNORE all other players on the field

These {len(frames_pil)} images are sequential frames at timestamps: {ts_str}

Identify these football actions performed ONLY by the RED BOX player:
• PASS   → kicks or passes ball to a teammate
• TACKLE → sliding or standing tackle / defensive challenge
• HEADER → contacts ball with their head
• SHOT   → shoots toward the goal

RULES:
- Only report actions where you are ≥75% confident
- One action per clear event (don't double-count the same action)
- If no action is visible from the RED BOX player, return empty list

Respond with ONLY valid JSON, no other text:
{{
  "actions": [
    {{"type": "PASS|TACKLE|HEADER|SHOT", "timestamp": <seconds>, "confidence": <0.0-1.0>}}
  ]
}}"""

    content = []
    for i, pil_img in enumerate(frames_pil):
        content.append(f"Frame at {timestamps[i]:.1f}s:")
        content.append(pil_img)
    content.append(prompt)

    for attempt in range(max_retries):
        try:
            response = model.generate_content(
                content,
                generation_config={"temperature": 0.1, "max_output_tokens": 1024},
            )
            text = response.text.strip()
            start = text.find("{")
            end = text.rfind("}") + 1
            if start >= 0 and end > start:
                data = json.loads(text[start:end])
                return data.get("actions", [])
            return []
        except Exception as e:
            err_str = str(e)
            if "429" in err_str or "quota" in err_str.lower() or "rate" in err_str.lower():
                wait = 20 * (2 ** attempt)
                print(f"  [rate limit] waiting {wait}s before retry {attempt+1}/{max_retries}...")
                time.sleep(wait)
            else:
                print(f"  [Gemini error] {e}")
                return []

    print("  [Gemini] max retries exceeded for this batch, skipping.")
    return []


def deduplicate_actions(actions: list, min_gap: float = 1.5) -> list:
    """Remove duplicate detections of the same action within min_gap seconds."""
    if not actions:
        return []
    sorted_a = sorted(actions, key=lambda x: x["timestamp"])
    result = [sorted_a[0]]
    for action in sorted_a[1:]:
        last = result[-1]
        if action["type"] == last["type"] and (action["timestamp"] - last["timestamp"]) < min_gap:
            # Keep the one with higher confidence
            if action.get("confidence", 0) > last.get("confidence", 0):
                result[-1] = action
        else:
            result.append(action)
    return result


# ──────────────────────────────────────────────
# Background analysis task
# ──────────────────────────────────────────────
def run_analysis(session_id: str, bbox_init: tuple, api_key: str):
    session = sessions[session_id]
    video_path = session["video_path"]

    try:
        session["status"] = "tracking"
        session["progress"] = 5

        # ── 1. Open video ──────────────────────────────────────
        cap = cv2.VideoCapture(video_path)
        fps = cap.get(cv2.CAP_PROP_FPS) or 25.0
        total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))

        # Read first frame (already consumed at upload, seek back)
        cap.set(cv2.CAP_PROP_POS_FRAMES, 0)
        ret, first_frame = cap.read()
        if not ret:
            raise RuntimeError("Cannot read first frame")

        # ── 2. Initialize tracker ─────────────────────────────
        tracker = create_tracker()
        tracker.init(first_frame, bbox_init)

        # Sample at 2 fps
        # 1fps sampling — enough for action detection, minimizes API calls
        sample_every = max(1, int(fps))
        tracked_data = []  # [{timestamp, pil_image, bbox}]

        frame_idx = 0
        last_valid_bbox = bbox_init
        prev_center = (
            bbox_init[0] + bbox_init[2] / 2,
            bbox_init[1] + bbox_init[3] / 2
        )
        max_jump_ratio = 0.25  # max jump relative to frame width

        frame_w = first_frame.shape[1]

        while True:
            ret, frame = cap.read()
            if not ret:
                break

            frame_idx += 1

            if frame_idx % sample_every != 0:
                continue

            # Update tracker
            success, bbox = tracker.update(frame)
            timestamp = frame_idx / fps

            if success and bbox_is_valid(bbox, frame.shape):
                cx = bbox[0] + bbox[2] / 2
                cy = bbox[1] + bbox[3] / 2
                jump = ((cx - prev_center[0]) ** 2 + (cy - prev_center[1]) ** 2) ** 0.5

                # Reject implausibly large jumps (tracker drift)
                if jump < max_jump_ratio * frame_w:
                    annotated = draw_player_box(frame, bbox)
                    pil_img = cv2_to_pil(annotated)
                    # Resize to save memory / API quota
                    pil_img = pil_img.resize((640, int(640 * pil_img.height / pil_img.width)))
                    tracked_data.append({
                        "timestamp": timestamp,
                        "pil": pil_img,
                        "bbox": bbox
                    })
                    last_valid_bbox = bbox
                    prev_center = (cx, cy)

            # Update progress (tracking phase = 5-50%)
            pct = 5 + int(45 * frame_idx / max(total_frames, 1))
            session["progress"] = min(pct, 50)

        cap.release()

        session["frames_tracked"] = len(tracked_data)
        session["status"] = "analyzing"
        session["progress"] = 50

        if not tracked_data:
            raise RuntimeError("Tracker lost the player immediately. Please select a clearer bounding box.")

        # ── 3. Gemini analysis ───────────────────────────────
        batch_size = 8
        all_actions = []
        total_batches = (len(tracked_data) + batch_size - 1) // batch_size

        for b_idx in range(0, len(tracked_data), batch_size):
            batch = tracked_data[b_idx: b_idx + batch_size]
            frames_pil = [item["pil"] for item in batch]
            timestamps = [item["timestamp"] for item in batch]

            actions = gemini_analyze_batch(api_key, frames_pil, timestamps)
            all_actions.extend(actions)

            done_batches = (b_idx // batch_size) + 1
            pct = 50 + int(40 * done_batches / max(total_batches, 1))
            session["progress"] = min(pct, 90)

            if b_idx + batch_size < len(tracked_data):
                time.sleep(4)

        # ── 4. Post-process ───────────────────────────────────
        valid = [
            a for a in all_actions
            if a.get("type", "").upper() in ("PASS", "TACKLE", "HEADER", "SHOT")
            and a.get("confidence", 0) >= 0.75
        ]
        deduped = deduplicate_actions(valid, min_gap=1.5)

        counts = {"PASS": 0, "TACKLE": 0, "HEADER": 0, "SHOT": 0}
        for a in deduped:
            counts[a["type"].upper()] += 1

        video_duration = tracked_data[-1]["timestamp"] if tracked_data else 0

        session["analysis"] = {
            "counts": counts,
            "timeline": sorted(deduped, key=lambda x: x["timestamp"]),
            "frames_analyzed": len(tracked_data),
            "video_duration": round(video_duration, 1),
        }
        session["status"] = "done"
        session["progress"] = 100

    except Exception as e:
        session["status"] = "error"
        session["error"] = str(e)
        print(f"[Analysis error] {e}")


# ──────────────────────────────────────────────
# Endpoints
# ──────────────────────────────────────────────
@app.post("/validate-key")
async def validate_key(req: ValidateKeyRequest):
    try:
        genai.configure(api_key=req.api_key)
        model = genai.GenerativeModel("gemini-2.0-flash")
        model.generate_content("ok")
        return {"ok": True}
    except Exception as e:
        raise HTTPException(status_code=401, detail=str(e))


@app.post("/upload")
async def upload_video(file: UploadFile = File(...)):
    session_id = str(uuid.uuid4())
    suffix = Path(file.filename).suffix or ".mp4"
    video_path = UPLOAD_DIR / f"{session_id}{suffix}"

    content = await file.read()
    with open(video_path, "wb") as f:
        f.write(content)

    cap = cv2.VideoCapture(str(video_path))
    ret, frame = cap.read()
    fps = cap.get(cv2.CAP_PROP_FPS) or 25.0
    total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    cap.release()

    if not ret:
        raise HTTPException(status_code=400, detail="No se pudo leer el video")

    sessions[session_id] = {
        "video_path": str(video_path),
        "status": "idle",
        "progress": 0,
        "analysis": None,
        "error": None,
    }

    return {
        "session_id": session_id,
        "first_frame": frame_to_b64(frame, quality=85),
        "width": frame.shape[1],
        "height": frame.shape[0],
        "fps": fps,
        "duration": round(total_frames / fps, 1),
    }


@app.post("/analyze")
async def analyze(req: AnalyzeRequest, background_tasks: BackgroundTasks):
    if req.session_id not in sessions:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")

    session = sessions[req.session_id]
    if session["status"] in ("tracking", "analyzing"):
        raise HTTPException(status_code=409, detail="Análisis ya en progreso")

    # Reset
    session.update({"status": "tracking", "progress": 0, "analysis": None, "error": None})

    bbox = (req.x, req.y, req.width, req.height)
    background_tasks.add_task(run_analysis, req.session_id, bbox, req.api_key)

    return {"ok": True, "session_id": req.session_id}


@app.get("/progress/{session_id}")
async def progress(session_id: str):
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")
    s = sessions[session_id]
    return {
        "status": s["status"],
        "progress": s["progress"],
        "frames_tracked": s.get("frames_tracked"),
        "error": s.get("error"),
        "analysis": s.get("analysis"),
    }


@app.get("/export/{session_id}")
async def export_excel(session_id: str):
    if session_id not in sessions:
        raise HTTPException(status_code=404, detail="Sesión no encontrada")

    analysis = sessions[session_id].get("analysis")
    if not analysis:
        raise HTTPException(status_code=400, detail="No hay análisis disponible")

    wb = openpyxl.Workbook()

    # ── Summary sheet ──────────────────────────────────────────
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False

    dark = "0D0D1A"
    green = "00FF88"
    mid = "1A1A2E"
    white = "FFFFFF"
    gray = "AAAAAA"

    def hdr(cell, text, bg=dark, fg=green, bold=True, size=11):
        ws[cell] = text
        ws[cell].font = Font(name="Calibri", color=fg, bold=bold, size=size)
        ws[cell].fill = PatternFill("solid", fgColor=bg)
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")

    def val(cell, text, bg=mid, fg=white, bold=False):
        ws[cell] = text
        ws[cell].font = Font(name="Calibri", color=fg, bold=bold, size=12)
        ws[cell].fill = PatternFill("solid", fgColor=bg)
        ws[cell].alignment = Alignment(horizontal="center", vertical="center")

    # Title
    ws.merge_cells("A1:D1")
    hdr("A1", "⚽ ANÁLISIS DE JUGADOR", dark, green, True, 16)
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:D2")
    hdr("A2", f"Video duration: {analysis['video_duration']}s  |  Frames analyzed: {analysis['frames_analyzed']}", mid, gray, False, 9)
    ws.row_dimensions[2].height = 18

    # Headers
    for col, label in enumerate(["ACCIÓN", "CANTIDAD", "PORCENTAJE", "DETALLE"], start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = label
        cell.font = Font(name="Calibri", color=green, bold=True, size=11)
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 24

    counts = analysis["counts"]
    total = sum(counts.values()) or 1
    icons = {"PASS": "🟢 PASE", "TACKLE": "🔴 BARRIDA", "HEADER": "🔵 CABEZAZO", "SHOT": "🟡 DISPARO"}
    descriptions = {"PASS": "Pases completados", "TACKLE": "Barridas / tackles", "HEADER": "Cabezazos", "SHOT": "Disparos al arco"}

    for r, key in enumerate(["PASS", "TACKLE", "HEADER", "SHOT"], start=5):
        count = counts.get(key, 0)
        pct = f"{count / total * 100:.0f}%"
        row_bg = "111126" if r % 2 == 0 else mid
        for col, v in enumerate([icons[key], count, pct, descriptions[key]], start=1):
            c = ws.cell(row=r, column=col, value=v)
            c.font = Font(name="Calibri", color=white, size=12, bold=(col == 2))
            c.fill = PatternFill("solid", fgColor=row_bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[r].height = 22

    # Column widths
    for col, w in enumerate([22, 12, 14, 26], start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    # ── Timeline sheet ─────────────────────────────────────────
    ws2 = wb.create_sheet("Timeline")
    ws2.sheet_view.showGridLines = False

    for col, (label, w) in enumerate([("TIEMPO (s)", 14), ("ACCIÓN", 16), ("CONFIANZA", 14)], start=1):
        c = ws2.cell(row=1, column=col, value=label)
        c.font = Font(name="Calibri", color=green, bold=True, size=11)
        c.fill = PatternFill("solid", fgColor=dark)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 22

    for r, event in enumerate(analysis["timeline"], start=2):
        row_bg = "111126" if r % 2 == 0 else mid
        for col, v in enumerate([
            round(event["timestamp"], 1),
            event["type"],
            f"{event.get('confidence', 0) * 100:.0f}%"
        ], start=1):
            c = ws2.cell(row=r, column=col, value=v)
            c.font = Font(name="Calibri", color=white, size=11)
            c.fill = PatternFill("solid", fgColor=row_bg)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws2.row_dimensions[r].height = 20

    output_path = UPLOAD_DIR / f"report_{session_id}.xlsx"
    wb.save(output_path)

    return FileResponse(
        str(output_path),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="player_analysis.xlsx"
    )


# ── Static files (must be last) ────────────────────────────────
app.mount("/", StaticFiles(directory="static", html=True), name="static")
